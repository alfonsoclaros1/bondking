from datetime import date, timedelta, datetime
from django.utils import timezone
import os
import traceback
from urllib import request
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError, PermissionDenied
from django.http import HttpResponseForbidden, JsonResponse,HttpResponse, QueryDict
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth import get_user_model
from django.db.models import Prefetch, Sum, Q
from django.core.paginator import Paginator
import openpyxl
from openpyxl import Workbook
from django.contrib import messages
from django.db import transaction
import pdfkit
from django.template.loader import get_template,render_to_string
from django.templatetags.static import static

from bondking import settings

from .models import (
    DR_STEP_META,
    PO_FLOW,
    PO_META,
    ApprovalStatus,
    Billing,
    BillingStatus,
    Client,
    DeliveryMethod,
    DeliveryReceiptItem,
    DeliveryStatus,
    InventoryIssuanceItem,
    POApprovalStatus,
    POStatus,
    PaymentStatus,
    Product,
    DeliveryReceipt,
    KANBAN_COLUMNS,
    PaymentMethod,
    ProductID,
    can_cancel_po,
    can_manage_inventory_issuance,
    get_effective_role,
    get_user_role,
    is_top_management,
    PurchaseOrder,
    InventoryIssuance,
    InventoryIssuanceItem,
    DeliveryReceiptItem,
    Product,
)
from .forms import (
    BillingFormSet,
    ClientForm,
    DeliveryReceiptForm,
    DeliveryReceiptItemFormSet,
    InventoryIssuanceForm,
    InventoryIssuanceItemFormSet,
    PurchaseOrderForm,
    PurchaseOrderParticularFormSet,
)

User = get_user_model()


# =========================
#  API HELPERS (CLIENT / PRODUCT)
# =========================
def root_redirect(request):
    if request.user.is_authenticated:
        return redirect("dr-kanban")
    return redirect("login")
@require_GET
@login_required
def client_detail_api(request, pk):
    client = get_object_or_404(Client, pk=pk)
    data = {
        "id": client.id,
        "company_name": client.company_name,
        "name_of_owner": client.name_of_owner,
        "contact_number": client.contact_number,
        "full_address": client.full_address,
    }
    return JsonResponse(data)


@require_GET
@login_required
def product_detail_api(request, pk):
    product = get_object_or_404(Product, pk=pk)
    data = {
        "id": product.id,
        "sku": product.sku,
        "name": product.name,
        "default_unit_price": str(product.default_unit_price or ""),
        "unit": product.unit,
    }
    return JsonResponse(data)

@require_GET
@login_required
def dr_filter_suggestions_api(request):
    """
    Global DR table smart-search suggestions.
    Returns mixed suggestion types across ALL records (no pagination limits).
    """
    q = (request.GET.get("q") or "").strip()
    if not q:
        return JsonResponse({"ok": True, "results": []})

    q_lower = q.lower()

    results = []

    # ---- Agents (match username + full name) ----
    agent_qs = (
        User.objects
        .filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q)
        )
        .order_by("username")[:10]
    )
    for u in agent_qs:
        label = u.get_full_name() or u.username
        results.append({
            "type": "agent",
            "badge": "Agent",
            "value": str(u.id),
            "label": label,
        })

    # ---- Clients (company name ONLY; return as name, not ID) ----
    client_qs = (
        Client.objects
        .filter(company_name__icontains=q)
        .order_by("company_name")[:10]
    )
    for c in client_qs:
        results.append({
            "type": "client_name",
            "badge": "Client",
            "value": c.company_name,   # ✅ NAME (not id)
            "label": c.company_name,
        })

    # ---- DR numbers (across ALL DRs) ----
    dr_qs = (
        DeliveryReceipt.objects
        .filter(dr_number__icontains=q)
        .order_by("dr_number")
        .values_list("dr_number", flat=True)[:10]
    )
    for drn in dr_qs:
        results.append({
            "type": "dr_number",
            "badge": "DR #",
            "value": drn,
            "label": drn,
        })

    # ---- Statuses (humanized) ----
    # PaymentStatus choices
    for val, label in PaymentStatus.choices:
        if q_lower in label.lower() or q_lower in val.lower():
            results.append({
                "type": "payment_status",
                "badge": "Payment Status",
                "value": val,
                "label": label,  # already humanized
            })

    # DeliveryStatus choices
    for val, label in DeliveryStatus.choices:
        if q_lower in label.lower() or q_lower in val.lower():
            results.append({
                "type": "delivery_status",
                "badge": "Delivery Status",
                "value": val,
                "label": label,
            })

    # PaymentMethod choices
    for val, label in PaymentMethod.choices:
        if q_lower in label.lower() or q_lower in val.lower():
            results.append({
                "type": "payment_method",
                "badge": "Payment Method",
                "value": val,
                "label": label,
            })

    # DeliveryMethod choices
    for val, label in DeliveryMethod.choices:
        if q_lower in label.lower() or q_lower in val.lower():
            results.append({
                "type": "delivery_method",
                "badge": "Delivery Method",
                "value": val,
                "label": label,
            })

    # ---- Free text fallback (maps to Client OR Agent) ----
    results.append({
        "type": "q",
        "badge": "Keyword",
        "value": q,
        "label": f'Search "{q}" (Client or Agent)',
    })

    # Keep results reasonable
    return JsonResponse({"ok": True, "results": results[:25]})

# =========================
#  CLIENT CREATE
# =========================

@login_required
def client_create(request):
    if request.method == "POST":
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save()
            return redirect("dr-create")  # or wherever you want
    else:
        form = ClientForm()
    return render(request, "bondking_app/client_form.html", {"form": form})


# =========================
#  DR CREATE / EDIT / DETAIL
# =========================

@login_required
def dr_create(request):
    """
    Create a new DeliveryReceipt with items.
    """
    DeliveryReceipt.get_d2d_stocks_client()

    if request.method == "POST":
        form = DeliveryReceiptForm(request.POST, request.FILES or None, stage="NEW_DR", user=request.user)
        formset = DeliveryReceiptItemFormSet(
            request.POST,
            prefix="items",
            stage="NEW_DR",
        )
        if form.is_valid() and formset.is_valid():
            dr = form.save(commit=False)
            dr.created_by = request.user
            dr.save()
            formset.instance = dr
            formset.save()
            return redirect("dr-kanban")
    else:
        initial = {
            "payment_method": PaymentMethod.CASH,
            "delivery_method": "DELIVERY",
            "agent": request.user.id,
        }
        form = DeliveryReceiptForm(initial=initial, stage="NEW_DR", user=request.user)
        formset = DeliveryReceiptItemFormSet(prefix="items", stage="NEW_DR")

    context = {
        "dr": None,
        "form": form,
        "formset": formset,
        "stage": "NEW_DR",
        "is_create": True,
        "updates": [],   # ✔ safe for create mode
        "clients": Client.objects.all().order_by("company_name"),
    }
    return render(request, "bondking_app/dr_form.html", context)

@login_required
def dr_edit(request, pk):
    dr = get_object_or_404(DeliveryReceipt, pk=pk)
    stage = dr.get_current_column()
    updates = dr.updates.all()
    action = None


    role = get_user_role(request.user)
    is_super = request.user.is_superuser

    current_step, next_step = dr.get_current_and_next_step()
    lifecycle_steps = dr.get_lifecycle_steps()
    current_meta = DR_STEP_META.get(current_step, {})
    next_meta = dr.get_next_step_meta()
    # =========================
    # DR NAVIGATION (Prev / Next)
    # =========================
    nav_qs = (
        DeliveryReceipt.objects
        .filter(is_archived=False, is_cancelled=False)
        .order_by("id")
    )

    prev_dr = None
    next_dr = None

    nav_from = request.GET.get("from")
    nav_ids = request.GET.get("nav_ids")

    # Keep full querystring EXCEPT page (so prev/next keeps context)
    qs = request.GET.copy()
    qs.pop("page", None)
    nav_querystring = qs.urlencode()
    # =========================
    # DR NAVIGATION (nav_ids OVERRIDE — NON-DESTRUCTIVE)
    # =========================
    prev_dr = None
    next_dr = None

    nav_ids = request.GET.get("nav_ids")

    if nav_ids:
        try:
            id_list = [int(x) for x in nav_ids.split(",")]

            if dr.id in id_list:
                idx = id_list.index(dr.id)

                if idx > 0:
                    prev_dr = DeliveryReceipt.objects.filter(id=id_list[idx - 1]).first()

                if idx < len(id_list) - 1:
                    next_dr = DeliveryReceipt.objects.filter(id=id_list[idx + 1]).first()
        except Exception:
            pass
    # -------------------------------
    # CASE 1: FROM KANBAN
    # -------------------------------
    if nav_from == "kanban" and nav_ids:
        try:
            id_list = [int(x) for x in nav_ids.split(",")]
            if dr.id in id_list:
                idx = id_list.index(dr.id)
                if idx > 0:
                    prev_dr = DeliveryReceipt.objects.filter(id=id_list[idx - 1]).first()
                if idx < len(id_list) - 1:
                    next_dr = DeliveryReceipt.objects.filter(id=id_list[idx + 1]).first()
        except Exception:
            pass

    # -------------------------------
    # CASE 2: FROM TABLE
    # -------------------------------
    elif nav_from == "table" and not nav_ids:
        # IMPORTANT: reuse the SAME queryset logic as dr_table
        qs_table = DeliveryReceipt.objects.all()

        if request.GET.get("hide_archived"):
            qs_table = qs_table.filter(is_archived=False)

        if request.GET.get("hide_cancelled"):
            qs_table = qs_table.filter(is_cancelled=False)

        if request.GET.get("payment_status"):
            qs_table = qs_table.filter(payment_status=request.GET["payment_status"])

        if request.GET.get("delivery_status"):
            qs_table = qs_table.filter(delivery_status=request.GET["delivery_status"])

        if request.GET.get("payment_method"):
            qs_table = qs_table.filter(payment_method=request.GET["payment_method"])

        if request.GET.get("delivery_method"):
            qs_table = qs_table.filter(delivery_method=request.GET["delivery_method"])

        if request.GET.get("client_name"):
            qs_table = qs_table.filter(client__company_name__icontains=request.GET["client_name"])

        if request.GET.get("dr_number"):
            qs_table = qs_table.filter(dr_number__icontains=request.GET["dr_number"])

        if request.GET.get("with_sales_invoice"):
            qs_table = qs_table.exclude(sales_invoice_no__isnull=True).exclude(sales_invoice_no="")

        # ---- SORTING (MATCH dr_table EXACTLY)
        sort_by = request.GET.get("sort_by", "date_desc")

        if sort_by == "date_asc":
            qs_table = qs_table.order_by("date_of_order")
        elif sort_by == "total_desc":
            qs_table = qs_table.order_by("-total_amount")
        elif sort_by == "total_asc":
            qs_table = qs_table.order_by("total_amount")
        elif sort_by == "dr_desc":
            qs_table = qs_table.order_by("-dr_number")
        elif sort_by == "dr_asc":
            qs_table = qs_table.order_by("dr_number")
        else:
            qs_table = qs_table.order_by("-date_of_order")

        ids = list(qs_table.values_list("id", flat=True))

        if dr.id in ids:
            idx = ids.index(dr.id)
            if idx > 0:
                prev_dr = qs_table.filter(id=ids[idx - 1]).first()
            if idx < len(ids) - 1:
                next_dr = qs_table.filter(id=ids[idx + 1]).first()
    # ==========================
    # ACTION HANDLING (PO STYLE)
    # ==========================
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "approve":
            dr.approve_current_step(request.user)
            return redirect("dr-edit", pk=dr.pk)

        if action == "decline":
            dr.decline_current_step(request.user)
            return redirect("dr-edit", pk=dr.pk)
        if action == "archive":
            if not (request.user.is_superuser or request.user.groups.filter(name="TopManagement").exists()):
                messages.error(request, "Only Top Management can archive DRs.")
                return redirect("dr-edit", pk=dr.pk)

            # Sample can archive at Delivered; others require Deposited
            can_archive = (
                (dr.delivery_method == DeliveryMethod.SAMPLE and dr.delivery_status == DeliveryStatus.DELIVERED)
                or (dr.payment_status == PaymentStatus.DEPOSITED)
            )
            if not can_archive:
                messages.error(request, "This DR is not yet eligible for archiving.")
                return redirect("dr-edit", pk=dr.pk)


            dr.is_archived = True
            dr.save(update_fields=["is_archived", "updated_at"])
            dr.log_update(request.user, "Delivery Receipt was archived.")
            messages.success(request, "Delivery Receipt archived successfully.")
            return redirect("dr-edit", pk=dr.pk)

    # ==========================
    # FORMS (SAVE CHANGES ONLY)
    # ==========================
    form = DeliveryReceiptForm(
        request.POST or None,
        request.FILES or None,
        instance=dr,
        stage=stage,
        user=request.user,
    )

    formset = DeliveryReceiptItemFormSet(
        request.POST or None,
        instance=dr,
        prefix="items",
        stage=stage,
    )

    if request.method == "POST" and request.POST.get("action") == "save":
        if not form.is_valid():
            print("DR FORM ERRORS:", form.errors)
        if not formset.is_valid():
            print("DR FORMSET ERRORS:", formset.errors)
        if form.is_valid() and formset.is_valid():
            old = DeliveryReceipt.objects.get(pk=dr.pk)

            dr = form.save(commit=False)
            formset.save()
            dr.save()


            # ---- FIELD CHANGE LOGGING ----
            for field in [
                "date_of_delivery",
                "payment_due",
                "payment_details",
                "remarks",
            ]:
                old_val = getattr(old, field)
                new_val = getattr(dr, field)
                if old_val != new_val:
                    label = field.replace("_", " ").title()
                    message = dr.log_update(
                        request.user,
                        f"{label} was set to {new_val} by {request.user.get_full_name() or request.user.username}",
                    )

            return redirect("dr-edit", pk=dr.pk)


    # ==========================
    # UI PERMISSIONS (SINGLE SOURCE)
    # ==========================
    can_approve = (
        dr.approval_status == ApprovalStatus.PENDING
        and (is_super or role in current_meta.get("approver_roles", set()))
    )

    can_decline = (
        dr.approval_status == ApprovalStatus.PENDING
        and (is_super or role in current_meta.get("decliner_roles", set()))
    )

    can_submit = (
        dr.approval_status == ApprovalStatus.APPROVED
        and next_step
        and (is_super or role in current_meta.get("forward_roles", set()))
    )


    lifecycle_steps = []
    steps = dr.get_lifecycle_steps()
    current = dr.get_current_column()
    current_index = steps.index(current) if current in steps else -1

    for idx, s in enumerate(steps):
        lifecycle_steps.append({
            "key": s,
            "label": s.replace("_", " ").title(),
            "is_current": s == current,
            "is_done": idx < current_index if current_index >= 0 else False,
        })

    current_step, next_step = dr.get_current_and_next_step()
    missing_fields = dr.get_missing_required_before_forward()


    kanban_url = reverse("dr-kanban")  # adjust name if needed

    if action == "resolve":
        if dr.approval_status != ApprovalStatus.DECLINED:
            messages.error(request, "This DR is not rejected.")
            return redirect("dr-edit", pk=dr.pk)

        resolved_note = (request.POST.get("resolved_note") or "").strip()
        if not resolved_note:
            messages.error(request, "Resolution note is required.")
            return redirect("dr-edit", pk=dr.pk)

        # Permission: only roles that can move this step
        meta = DR_STEP_META.get(dr.get_current_column(), {})
        allowed_roles = meta.get("forward_roles", set())
        role = get_user_role(request.user)

        if role not in allowed_roles and not request.user.is_superuser:
            messages.error(request, "You are not allowed to resolve this DR.")
            return redirect("dr-edit", pk=dr.pk)

        dr.reject_solution = resolved_note
        dr.approval_status = ApprovalStatus.PENDING
        dr.save(update_fields=["reject_solution", "approval_status", "updated_at"])

        dr.log_update(
            request.user,
            "Resolved rejection and returned DR to Pending approval.",
            user_notes=resolved_note,
        )
        return redirect("dr-edit", pk=dr.pk)


    # ==========================
    # CONTEXT
    # ==========================
    context = {
        "dr": dr,
        "form": form,
        "formset": formset,
        "updates": updates,

        "stage": stage,
        "is_create": False,

        "lifecycle_steps": lifecycle_steps,
        "current_step": current_step,
        "next_step": next_step,
        "next_meta": next_meta,
        "approval_status": dr.approval_status,
        "missing_fields": missing_fields,
        "can_approve": can_approve,
        "can_decline": can_decline,
        "kanban_url": kanban_url,
        "dr_flow": lifecycle_steps,
        "has_missing_required": bool(missing_fields),
        "is_top_management": is_top_management(request.user),
        "clients": Client.objects.all().order_by("company_name"),
        "prev_dr": prev_dr,
        "next_dr": next_dr,
        "nav_querystring": nav_querystring,

    }

    return render(request, "bondking_app/dr_form.html", context)


@login_required
def dr_detail(request, pk):
    dr = get_object_or_404(DeliveryReceipt, pk=pk)
    return render(request, "bondking_app/dr_detail.html", {"dr": dr})


# =========================
#  KANBAN BOARD
# =========================
@login_required
def dr_kanban(request):
    """
    Render the DR Kanban board.
    """
    active_drs = (
    DeliveryReceipt.objects
    .filter(is_archived=False, is_cancelled=False)
    .select_related("client", "agent")
    .order_by("-created_at")
    )

    normal_drs = active_drs.exclude(
        delivery_method=DeliveryMethod.D2D_STOCKS
    )

    d2d_stocks = active_drs.filter(
        delivery_method=DeliveryMethod.D2D_STOCKS
    )
    column_items = {col: [] for col in KANBAN_COLUMNS}

    user = request.user
    role = get_user_role(user)
    is_super = user.is_superuser
    top_mgmt = is_top_management(user)

    for dr in normal_drs:
        current_step = dr.get_current_column()
        step_meta = DR_STEP_META.get(current_step, {})

        dr.can_approve = (
            dr.approval_status == ApprovalStatus.PENDING
            and (
                is_super
                or top_mgmt
                or role in step_meta.get("approver_roles", set())
            )
        )

        dr.can_decline = (
            dr.approval_status == ApprovalStatus.PENDING
            and (
                is_super
                or top_mgmt
                or role in step_meta.get("decliner_roles", set())
            )
        )

        column_items[current_step].append(dr)




    d2d_stocks = DeliveryReceipt.objects.filter(
        delivery_method=DeliveryMethod.D2D_STOCKS,
        is_archived=False
    )



    # Labels for each column
    column_labels = {
        "NEW_DR": "New DR",
        "FOR_DELIVERY": "For Delivery",
        "DELIVERED": "Delivered",
        "FOR_COUNTER_CREATION": "For Counter Creation",
        "FOR_COUNTERING": "For Countering",
        "COUNTERED": "Countered",
        "FOR_COLLECTION": "For Collection",
        "FOR_DEPOSIT": "For Deposit",
        "DEPOSITED": "Deposited",
    }

    # Build render-friendly list to avoid any dict indexing in the template
    columns_render = []
    for key in KANBAN_COLUMNS:
        columns_render.append(
            {
                "key": key,
                "label": column_labels.get(key, key),
                "items": column_items.get(key, []),
            }
        )

    user = request.user
    is_admin_like = user.is_superuser 
    role = get_user_role(user)

    available_roles = [
        ("SalesAgent", "Sales Agent"),
        ("SalesHead", "Sales Head"),
        ("LogisticsOfficer", "Logistics Officer"),
        ("LogisticsHead", "Logistics Head"),
        ("AccountingOfficer", "Accounting Officer"),
        ("AccountingHead", "Accounting Head"),
        ("TopManagement", "Top Management"),
    ]
    top_mgmt = is_top_management(user)

    context = {
        "columns_render": columns_render,
        "is_admin_like": is_admin_like,
        "available_roles": available_roles if is_admin_like else [],
        "d2d_stocks": d2d_stocks,
        "column_items": column_items,
        "KANBAN_COLUMNS": KANBAN_COLUMNS,
        "is_top_management": top_mgmt,
    }
    return render(request, "bondking_app/dr_kanban.html", context)
# =========================
#  KANBAN ACTIONS (MOVE / APPROVE / DECLINE)
# =========================

@require_POST
@login_required
def move_dr(request, pk):
    dr = get_object_or_404(DeliveryReceipt, pk=pk)

    target_column = request.POST.get("target_column")
    notes = request.POST.get("notes", "")
    sim_role = request.POST.get("sim_role") or None
    # ==========================
    # ENFORCE REQUIRED FIELDS
    # ==========================
    step_meta = DR_STEP_META.get(target_column, {})
    required_fields = step_meta.get("required_fields", [])

    for field_name in required_fields:
        value = getattr(dr, field_name, None)

        if not value:
            verbose = dr._meta.get_field(field_name).verbose_name
            return JsonResponse({
                "ok": False,
                "error": f"{verbose} is required before proceeding to {step_meta.get('label', target_column)}."
            }, status=400)
    try:
        dr.move_to_column(
            request.user,
            target_column,
            user_notes=notes,
            simulated_role=sim_role,
        )
    except (ValidationError, PermissionDenied) as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    except ValidationError as e:
        return JsonResponse({
            "ok": False,
            "error": str(e),
        }, status=400)

    except PermissionDenied as e:
        return JsonResponse({
            "ok": False,
            "error": "You do not have permission to perform this action.",
        }, status=403)

    except Exception as e:
        # 👇 THIS IS THE KEY FIX
        traceback.print_exc()  # shows real error in console

        return JsonResponse({
            "ok": False,
            "error": str(e),
        }, status=500)

    # Recalculate counts per column for the board
    counts = {col: 0 for col in KANBAN_COLUMNS}
    for d in DeliveryReceipt.objects.all():
        col = d.get_current_column()
        counts[col] = counts.get(col, 0) + 1

    latest = dr.updates.first()
    system_message = latest.system_update if latest else ""

    return JsonResponse(
        {
            "ok": True,
            "new_column": dr.get_current_column(),
            "approval_status": dr.approval_status,
            "delivery_status": dr.delivery_status,
            "payment_status": dr.payment_status,
            "payment_status_display": dr.get_payment_status_display(),
            "payment_method_display": dr.get_payment_method_display(),
            "counts": counts,
            "system_message": system_message,  # ✅ ADD
        }
    )




@require_POST
@login_required
def dr_approve(request, pk):

    # ✅ ADD THIS (TOP OF FUNCTION)
    if request.headers.get("x-requested-with") != "XMLHttpRequest":
        return JsonResponse(
            {"ok": False, "error": "Invalid request type"},
            status=400
        )

    dr = get_object_or_404(DeliveryReceipt, pk=pk)

    notes = request.POST.get("notes", "")
    sim_role = request.POST.get("sim_role") or None

    try:
        dr.approve_current_step(
            request.user,
            user_notes=notes,
            simulated_role=sim_role,
        )

        current_stage = dr.get_current_column()
        if current_stage == "FOR_DELIVERY" and not dr.payment_due:
            dr.payment_due = date.today() + timedelta(days=3)
            dr.save(update_fields=["payment_due"])

    except (ValidationError, PermissionDenied) as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)

    except Exception as e:
        # ✅ PREVENT HTML ERROR PAGE
        traceback.print_exc()
        return JsonResponse({"ok": False, "error": str(e)}, status=500)
    latest = dr.updates.first()

    return JsonResponse({
        "ok": True,
        "approval_status": dr.approval_status,
        "system_message": latest.system_update if latest else "",  # ✅ ADD
    })



@require_POST
@login_required
def dr_decline(request, pk):
    dr = get_object_or_404(DeliveryReceipt, pk=pk)

    sim_role = request.POST.get("sim_role") or None

    # First attempt the decline
    try:
        reject_problem = (request.POST.get("reject_problem") or "").strip()
        reject_solution = (request.POST.get("reject_solution") or "").strip()

        if not reject_problem:
            return JsonResponse(
                {"ok": False, "error": "Rejection reason is required."},
                status=400,
            )

        dr.reject_problem = reject_problem
        dr.reject_solution = reject_solution

        dr.decline_current_step(
            request.user,
            user_notes=reject_problem,
            simulated_role=sim_role,
        )
        dr.save(update_fields=["reject_problem", "reject_solution", "approval_status", "updated_at"])

    except (ValidationError, PermissionDenied) as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)

    # ============================================================
    # SPECIAL RULE FOR CASH DRs:
    # If declined while in FOR_DEPOSIT → move back to DELIVERED
    # ============================================================
    try:
        pm = str(dr.payment_method).upper()
        current_column = dr.get_current_column()

        if pm == "CASH" and current_column == "FOR_DEPOSIT":
            dr.move_to_column(
                user=request.user,
                target_column="DELIVERED",
                user_notes="Declined at For Deposit – Auto-reverted to Delivered (Cash rule)",
            )
    except (ValidationError, PermissionDenied) as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    latest = dr.updates.first()

    # Return final state to JS
    return JsonResponse(
        {
            "ok": True,
            "approval_status": dr.approval_status,
            "column": dr.get_current_column(),
            "system_message": latest.system_update if latest else "",  # ✅ ADD
        }
    )

@require_POST
@login_required
def archive_dr(request, dr_id):
    if not request.user.groups.filter(name="TopManagement").exists() and not request.user.is_superuser:
        raise PermissionDenied("Only Top Management can archive DRs.")

    dr = get_object_or_404(DeliveryReceipt, pk=dr_id)

    can_archive = (
        (dr.delivery_method == DeliveryMethod.SAMPLE and dr.delivery_status == DeliveryStatus.DELIVERED)
        or (dr.payment_status == PaymentStatus.DEPOSITED)
    )
    if not can_archive:
        return JsonResponse({"ok": False, "error": "This DR is not yet eligible for archiving."}, status=400)

    dr.is_archived = True
    dr.save(update_fields=["is_archived", "updated_at"])
    dr.log_update(request.user, "Archived DR.")
    latest = dr.updates.first()
    return JsonResponse({"ok": True, "system_message": latest.system_update if latest else ""})



@require_GET
def d2d_transactions_api(request, pk):
    """
    Returns per-product breakdown of d2d Stocks usage
    """
    dr = get_object_or_404(
        DeliveryReceipt,
        pk=pk,
        delivery_method=DeliveryMethod.D2D_STOCKS,
        is_archived=False,
    )

    # Build item breakdown
    result = []

    for item in dr.items.select_related("product"):
        issued_qty = item.quantity

        d2d_items = (
            DeliveryReceiptItem.objects
            .filter(
                delivery_receipt__delivery_method=DeliveryMethod.DOOR_TO_DOOR,
                delivery_receipt__source_dr=dr,
                delivery_receipt__is_cancelled=False,
                product=item.product,
            )
            .select_related("delivery_receipt", "delivery_receipt__client")
        )


        used_qty = sum(i.quantity for i in d2d_items)
        remaining_qty = issued_qty - used_qty

        transactions = []
        for i in d2d_items:
            transactions.append({
                "dr_number": i.delivery_receipt.dr_number,
                "client": i.delivery_receipt.client.company_name if i.delivery_receipt.client else "—",
                "date": i.delivery_receipt.date_of_order.strftime("%Y-%m-%d"),
                "quantity": i.quantity,
            })

        result.append({
            # ---- AUTOFILL DATA ----
            "product_id": item.product_id,
            "product_name": item.product.name,
            "remaining_qty": max(remaining_qty, 0),
            "unit_price": float(item.unit_price),
            "description": item.description or "",

            # ---- REPORTING DATA ----
            "issued_qty": issued_qty,
            "used_qty": used_qty,
            "transactions": transactions,
        })


    can_archive = all(r["remaining_qty"] <= 0 for r in result)

    return JsonResponse({
        "ok": True,
        "dr_number": dr.dr_number,
        "agent": dr.agent.get_full_name() or dr.agent.username,
        "items": result,
        "can_archive": can_archive,
        "is_top_management": request.user.is_superuser,
    })
def clean_int_list(values):
    out = []
    for v in values:
        if v in (None, "", "None", "null", "undefined"):
            continue
        if str(v).isdigit():
            out.append(int(v))
    return out

def clean_int(val):
    """
    Safely convert GET param to int.
    Returns None if empty, 'None', non-numeric, etc.
    """
    if val in (None, "", "None", "null", "undefined"):
        return None
    if str(val).isdigit():
        return int(val)
    return None
def clean_param(val):
    """
    Returns None if the param is empty, 'None', 'null', etc.
    """
    if val in (None, "", "None", "null", "undefined"):
        return None
    return val

def excel_safe_datetime(value):
    """
    Ensures Excel-friendly datetime/date output.
    - Converts aware datetimes to naive (localtime)
    - Leaves strings/numbers untouched
    """
    if value is None:
        return ""

    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.replace(tzinfo=None)

    if isinstance(value, date):
        return value

    return value

@login_required
def dr_table(request):
    qs = (
        DeliveryReceipt.objects
        .select_related("client", "agent")
        .prefetch_related("items__product")
    )

    # -------------------
    # Show All logic
    # -------------------
    def clean_param(val):
        """
        Returns None if the param is empty, 'None', 'null', etc.
        """
        if val in (None, "", "None", "null", "undefined"):
            return None
        return val


    # ---- Smart search tags (multi) ----
    agent_ids = clean_int_list(request.GET.getlist("agent"))
    client_names = [c.strip() for c in request.GET.getlist("client_name") if c.strip()]  # ✅ name-based only
    dr_numbers = [d.strip() for d in request.GET.getlist("dr_number") if d.strip()]

    payment_methods = [p for p in request.GET.getlist("payment_method") if p]
    payment_statuses = [p for p in request.GET.getlist("payment_status") if p]
    delivery_statuses = [d for d in request.GET.getlist("delivery_status") if d]
    delivery_methods = [d for d in request.GET.getlist("delivery_method") if d]
    
    # ---- Free-text keyword (client OR agent) ----
    q = (request.GET.get("q") or "").strip()
    # -------------------
    # Filters
    # -------------------

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    sort_by = request.GET.get("sort_by", "dr_asc")
    due_start = request.GET.get("due_start", "")
    due_end = request.GET.get("due_end", "")
    hide_archived = request.GET.get("hide_archived") == "1"
    hide_cancelled = request.GET.get("hide_cancelled") == "1"
    with_sales_invoice = request.GET.get("with_sales_invoice") == "1"


    if hide_archived:
        qs = qs.filter(is_archived=False)

    if hide_cancelled:
        qs = qs.filter(is_cancelled=False)


    if with_sales_invoice:
        qs = qs.exclude(
            sales_invoice_no__isnull=True
        ).exclude(
            sales_invoice_no__exact=""
        )

    # --- DR numbers (multi) ---
    if dr_numbers:
        q_obj = Q()
        for drn in dr_numbers:
            q_obj |= Q(dr_number__icontains=drn)
        qs = qs.filter(q_obj)


    if due_start:
        qs = qs.filter(payment_due__gte=due_start)

    if due_end:
        qs = qs.filter(payment_due__lte=due_end)

    # --- Delivery method (multi) ---
    if delivery_methods:
        qs = qs.filter(delivery_method__in=delivery_methods)

    # --- Agent (multi) ---
    if agent_ids:
        qs = qs.filter(agent_id__in=agent_ids)

    # --- Client name (multi, NAME-BASED ONLY) ---
    if client_names:
        q_obj = Q()
        for cname in client_names:
            q_obj |= Q(client__company_name__icontains=cname)
        qs = qs.filter(q_obj)

    # --- Payment/Delivery filters (multi) ---
    if payment_methods:
        qs = qs.filter(payment_method__in=payment_methods)

    if payment_statuses:
        qs = qs.filter(payment_status__in=payment_statuses)

    if delivery_statuses:
        qs = qs.filter(delivery_status__in=delivery_statuses)

    if start_date:
        qs = qs.filter(date_of_order__gte=start_date)

    if end_date:
        qs = qs.filter(date_of_order__lte=end_date)

    client = clean_param(request.GET.get("client"))

    if client and client.isdigit():
        qs = qs.filter(client_id=int(client))

    # --- Free-text keyword maps to Client OR Agent ---
    if q:
        qs = qs.filter(
            Q(client__company_name__icontains=q) |
            Q(agent__username__icontains=q) |
            Q(agent__first_name__icontains=q) |
            Q(agent__last_name__icontains=q)
        )
    SORT_OPTIONS = {
        "date_desc": "-date_of_order",
        "date_asc": "date_of_order",
        "total_desc": "-total_amount",
        "total_asc": "total_amount",
        "dr_desc": "-dr_number",
        "dr_asc": "dr_number",
    }
    qs = qs.order_by(SORT_OPTIONS.get(sort_by, "-date_of_order"))

    # ---- Client display value (for template only) ----
    client_display = ""

    if client and client.isdigit():
        try:
            client_display = Client.objects.get(pk=int(client)).company_name
        except Client.DoesNotExist:
            client_display = ""

    # -------------------
    # Pagination (LAST)
    # -------------------
    paginator = Paginator(qs, 100)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    client_display = ", ".join(client_names) if client_names else q
    total_sum = qs.aggregate(total=Sum("total_amount"))["total"] or 0

    context = {
        "page_obj": page_obj,
        "clients": Client.objects.all().order_by("company_name"),
        "agents": User.objects.all().order_by("username"),
        "payment_methods": PaymentMethod.choices,
        "payment_statuses": PaymentStatus.choices,
        "delivery_statuses": DeliveryStatus.choices,
        "hide_archived": hide_archived,
        "hide_cancelled": hide_cancelled,
        "sort_by": sort_by,
        "is_top_management": is_top_management(request.user),
        "delivery_methods": DeliveryMethod.choices,
        "client": client,
        "client_name": client_names,
        "client_display": client_display,
        "total_sum": total_sum,
        "selected": {
            "agent": agent_ids,
            "client_name": client_names,
            "dr_number": dr_numbers,
            "payment_method": payment_methods,
            "payment_status": payment_statuses,
            "delivery_status": delivery_statuses,
            "delivery_method": delivery_methods,
            "start_date": start_date,
            "end_date": end_date,
            "due_start": due_start,
            "due_end": due_end,
            "with_sales_invoice": with_sales_invoice,
            "q": q,
        },
    }

    return render(request, "bondking_app/dr_table.html", context)

@require_GET
@login_required
def dr_items_api(request, pk):
    from .models import DeliveryReceipt

    dr = get_object_or_404(
        DeliveryReceipt.objects.select_related("client", "agent").prefetch_related("items__product"),
        pk=pk,
    )

    items = []
    for it in dr.items.all():
        items.append({
            "product": it.product.name,
            "description": it.description or "",
            "quantity": float(it.quantity),
            "unit_price": float(it.unit_price),
            "line_total": float(it.line_total),
        })

    return JsonResponse({
        "ok": True,
        "dr_number": dr.dr_number,
        "items": items,
    })
@login_required
def dr_delete(request, pk):
    # Top Management only
    if not is_top_management(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Only Top Management can delete Delivery Receipts.")

    dr = get_object_or_404(DeliveryReceipt, pk=pk)

    if request.method != "POST":
        raise PermissionDenied("Invalid request method.")

    confirm_text = request.POST.get("confirm_text", "").strip()

    if confirm_text != dr.dr_number:
        messages.error(
            request,
            "Verification failed. Please type the exact DR Number to confirm deletion."
        )
        return redirect("dr-edit", pk=dr.pk)

    dr_number = dr.dr_number
    dr.delete()

    messages.success(
        request,
        f"Delivery Receipt {dr_number} was permanently deleted."
    )
    return redirect("dr-table")



@login_required
def po_create(request):
    # permission: Accounting Officer + Accounting Head only (plus superuser/top mgmt)
    role = get_user_role(request.user)
    if not (request.user.is_superuser or role in {"AccountingOfficer", "AccountingHead", "TopManagement"}):
        raise PermissionDenied("You are not allowed to create a PO.")

    if request.method == "POST":
        form = PurchaseOrderForm(request.POST, stage="PURCHASE_ORDER_CREATION", user=request.user)
        formset = PurchaseOrderParticularFormSet(request.POST, prefix="parts", stage="PURCHASE_ORDER_CREATION")
        if form.is_valid() and formset.is_valid():
            po = form.save(commit=False)
            po.prepared_by = request.user
            po.status = POStatus.PURCHASE_ORDER_APPROVAL
            po.approval_status = "PENDING"
            po.po_number = None
            po.save()
            formset.instance = po
            formset.save()
            return redirect("po-edit", pk=po.pk)
        if not form.is_valid():
            print("FORM ERRORS:", form.errors)

        if not formset.is_valid():
            print("FORMSET ERRORS:", formset.errors)

    else:
        form = PurchaseOrderForm(stage="PURCHASE_ORDER_CREATION", user=request.user)
        formset = PurchaseOrderParticularFormSet(prefix="parts", stage="PURCHASE_ORDER_CREATION")
    PO_FLOW = [
        POStatus.PURCHASE_ORDER_CREATION,
        POStatus.PURCHASE_ORDER_APPROVAL,
        POStatus.BILLING,
        POStatus.PO_FILING,
        POStatus.ARCHIVED,
    ]


    current_status = POStatus.PURCHASE_ORDER_CREATION

    po_flow = []
    for idx, code in enumerate(PO_FLOW):
        po_flow.append({
            "code": code,
            "label": code.replace("_", " ").title(),
            "is_current": code == current_status,
            "is_done": False,  # nothing completed yet
        })

    return render(request, "bondking_app/po_form.html", {
        "po": None,
        "form": form,
        "formset": formset,
        "stage": "PURCHASE_ORDER_CREATION",
        "is_create": True,
        "updates": [],
        "po_flow": po_flow,
        "can_edit": True,
        "can_submit": False,
        "can_approve": False,
        "can_archive": False,
        "next_actor": None,
        "next_step_description": "",
    })

@login_required
def po_edit(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    stage = po.status

    updates = po.updates.all() if hasattr(po, "updates") else []

    # ==========================
    # ACTION BUTTON HANDLING
    # ==========================
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "approve":
            po.approve_current_step(request.user)
            return redirect("po-edit", pk=po.pk)

        if action == "decline":
            po.decline_current_step(request.user)
            return redirect("po-table")

        if action == "submit":
            try:
                po.submit_to_next(request.user)
                return redirect("po-edit", pk=po.pk)
            except ValidationError as e:
                # Show popup error instead of crashing
                messages.error(request, e.messages[0] if hasattr(e, "messages") else str(e))
                
        if action == "archive":
            po.status = POStatus.ARCHIVED
            po.is_archived = True
            po.save(update_fields=["status", "is_archived", "updated_at"])
            po.log_update(request.user, "PO archived.")
            return redirect("po-table")

    # ==========================
    # FORMS (SAVE CHANGES ONLY)
    # ==========================
    form = PurchaseOrderForm(
        request.POST or None,
        instance=po,
        stage=stage,
        user=request.user,
    )

    formset = PurchaseOrderParticularFormSet(
        request.POST or None,
        instance=po,
        prefix="parts",
        stage=stage,
        approval_status=po.approval_status,
    )
    billing_formset = BillingFormSet(
        request.POST or None,
        instance=po,
        prefix="billings",
        stage=po.status,
        user=request.user,
    )
    print("=== VALIDATION DEBUG ===")
    print("PO form valid:", form.is_valid())
    print("PO errors:", form.errors)

    print("Particulars valid:", formset.is_valid())
    print("Particulars errors:", formset.errors)

    print("Billing valid:", billing_formset.is_valid())
    print("Billing errors:", billing_formset.errors)
    print("Billing non-form errors:", billing_formset.non_form_errors())
    if request.method == "POST" and request.POST.get("action") == "save":
        form = PurchaseOrderForm(
            request.POST or None,
            instance=po,
            stage=stage,
            user=request.user,
        )

        formset = PurchaseOrderParticularFormSet(
            request.POST or None,
            instance=po,
            prefix="parts",
            stage=stage,
            approval_status=po.approval_status,
        )
        billing_formset = BillingFormSet(
            request.POST or None,
            instance=po,
            prefix="billings",
            stage=po.status,
            user=request.user,
        )
        if form.is_valid() and formset.is_valid() and billing_formset.is_valid():
            po = form.save()
            formset.save()

            billings = billing_formset.save(commit=False)

            for billing in billings:
                if not billing.billing_number:
                    billing.billing_number = Billing.get_next_billing_number()
                    billing.source_po = po
                    billing.status = BillingStatus.CHECK_CREATION
                billing.save()

            billing_formset.save_m2m()
            return redirect("po-edit", po.id)

        if not form.is_valid():
            print("PO FORM ERRORS:", form.errors)

        if not formset.is_valid():
            print("PARTICULARS FORMSET ERRORS:", formset.errors)

        if not billing_formset.is_valid():
            print("BILLING FORMSET ERRORS:", billing_formset.errors)
            print("BILLING NON-FORM ERRORS:", billing_formset.non_form_errors())

    # ==========================
    # EDIT PERMISSIONS
    # ==========================
    can_edit = stage in [
        POStatus.PURCHASE_ORDER_CREATION,
        POStatus.BILLING,
    ]

    # ==========================
    # NEXT STEP CARD
    # ==========================
    NEXT_STEP_META = {
        POStatus.PURCHASE_ORDER_CREATION: ("RVT", "Accounting prepares and submits the Purchase Order."),
        POStatus.PURCHASE_ORDER_APPROVAL: ("JGG", "Purchase Order is awaiting approval."),
        POStatus.BILLING: ("RVT / Accounting", "Create billings (partial payments) until totals match and all billings are PAID."),
        POStatus.PO_FILING: ("RVT", "PO is ready to be archived."),
        POStatus.ARCHIVED: (None, "This Purchase Order has been archived."),
    }

    next_actor, next_step_description = NEXT_STEP_META.get(stage, (None, ""))

    # ==========================
    # PO LIFECYCLE (SINGLE SOURCE)
    # ==========================
    PO_FLOW = [
        POStatus.PURCHASE_ORDER_CREATION,
        POStatus.PURCHASE_ORDER_APPROVAL,
        POStatus.BILLING,
        POStatus.PO_FILING,
        POStatus.ARCHIVED,
    ]

    current_index = PO_FLOW.index(stage)

    po_flow = []
    for idx, code in enumerate(PO_FLOW):
        po_flow.append({
            "code": code,
            "label": code.replace("_", " ").title(),
            "is_current": code == stage,
            "is_done": idx < current_index,
        })
    SUBMIT_LABELS = {
        POStatus.PURCHASE_ORDER_CREATION: "Submit PO",
        POStatus.BILLING: "Proceed to PO Filing",
    }


    submit_label = SUBMIT_LABELS.get(stage, "Submit")

    # --------------------------
    # CAN SUBMIT LOGIC (FIXED)
    # --------------------------

    EARLY_SUBMIT_STAGES = {
        POStatus.PURCHASE_ORDER_CREATION,
    }

    APPROVAL_REQUIRED_SUBMIT_STAGES = {
        POStatus.BILLING,
    }

    role = get_user_role(request.user)
    meta = PO_META.get(stage, {})

    can_submit = (
        not po.is_archived
        and not po.is_cancelled
        and (request.user.is_superuser or role in meta.get("forward_roles", set()))
        and (not meta.get("requires_approval") or po.approval_status == POApprovalStatus.APPROVED)
    )

    # =========================
    # PO NAVIGATION (Prev / Next)
    # =========================
    prev_po = None
    next_po = None

    nav_from = request.GET.get("from")
    nav_ids = request.GET.get("nav_ids")
    qs = request.GET.copy()
    qs.pop("page", None)
    nav_querystring = qs.urlencode()
    if nav_from == "bulk" and nav_ids:
        try:
            ids = [int(i) for i in nav_ids.split(",")]
        except ValueError:
            ids = []

        if po.id in ids:
            idx = ids.index(po.id)
            if idx > 0:
                prev_po = PurchaseOrder.objects.filter(id=ids[idx - 1]).first()
            if idx < len(ids) - 1:
                next_po = PurchaseOrder.objects.filter(id=ids[idx + 1]).first()
    elif nav_from == "table":
        qs_table = PurchaseOrder.objects.all()

        # ---- APPLY SAME FILTERS AS po_table ----
        if request.GET.get("paid_to"):
            qs_table = qs_table.filter(paid_to__icontains=request.GET["paid_to"])

        if request.GET.get("prepared_by", "").isdigit():
            qs_table = qs_table.filter(prepared_by_id=int(request.GET["prepared_by"]))

        if request.GET.get("status"):
            qs_table = qs_table.filter(status=request.GET["status"])

        if request.GET.get("approval_status"):
            qs_table = qs_table.filter(approval_status=request.GET["approval_status"])

        if request.GET.get("start_date"):
            qs_table = qs_table.filter(date__gte=request.GET["start_date"])

        if request.GET.get("end_date"):
            qs_table = qs_table.filter(date__lte=request.GET["end_date"])

        if request.GET.get("product_id"):
            qs_table = qs_table.filter(product_id_ref_id=request.GET["product_id"])

        if request.GET.get("hide_archived"):
            qs_table = qs_table.filter(is_archived=False)

        if request.GET.get("hide_cancelled"):
            qs_table = qs_table.filter(is_cancelled=False)

        # ---- SORT (same keys as table) ----
        sort_by = request.GET.get("sort_by")
        if sort_by == "date_asc":
            qs_table = qs_table.order_by("date")
        elif sort_by == "date_desc":
            qs_table = qs_table.order_by("-date")
        elif sort_by == "total_asc":
            qs_table = qs_table.order_by("total")
        elif sort_by == "total_desc":
            qs_table = qs_table.order_by("-total")
        elif sort_by == "po_asc":
            qs_table = qs_table.order_by("po_number")
        elif sort_by == "po_desc":
            qs_table = qs_table.order_by("-po_number")
        else:
            qs_table = qs_table.order_by("-date")

        ids = list(qs_table.values_list("id", flat=True))

        if po.id in ids:
            idx = ids.index(po.id)
            if idx > 0:
                prev_po = PurchaseOrder.objects.filter(id=ids[idx - 1]).first()
            if idx < len(ids) - 1:
                next_po = PurchaseOrder.objects.filter(id=ids[idx + 1]).first()

    billed_total = po.billed_total()
    balance = po.balance_amount()
    return render(request, "bondking_app/po_form.html", {
        "po": po,
        "form": form,
        "formset": formset,
        "updates": updates,
        "can_edit": can_edit,
        "can_submit":can_submit,
        "can_approve": po.approval_status == POApprovalStatus.PENDING and stage in [
            POStatus.PURCHASE_ORDER_APPROVAL,
        ],
        "can_archive": stage == POStatus.PO_FILING,
        "next_actor": next_actor,
        "next_step_description": next_step_description,
        "po_flow": po_flow,
        "stage": stage,
        "is_create": False,
        "submit_label": submit_label,
        "billing_formset": billing_formset,
        "billed_total": billed_total,
        "balance": balance,
        "prev_po": prev_po,
        "next_po": next_po,
        "nav_querystring": nav_querystring,

    })



@require_POST
@login_required
def po_approve(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)

    notes = request.POST.get("notes", "")
    sim_role = request.POST.get("sim_role") or None

    try:
        po.approve_current_step(
            request.user,
            user_notes=notes,
            simulated_role=sim_role,
        )

    except (ValidationError, PermissionDenied) as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    latest = po.updates.first()
    return JsonResponse({
        "ok": True,
        "system_message": latest.system_update if latest else "",
    })


def is_attention_required(self):
    return self.approval_status == POApprovalStatus.DECLINED

@require_POST
@login_required
def po_decline(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    notes = request.POST.get("notes", "")
    sim_role = request.POST.get("sim_role") or None

    try:
        po.decline_current_step(request.user, user_notes=notes, simulated_role=sim_role)
    except (ValidationError, PermissionDenied) as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    latest = po.updates.first()
    return JsonResponse({
        "ok": True,
        "approval_status": po.approval_status,
        "column": po.get_current_column(),
        "system_message": latest.system_update if latest else "",
    })


@require_POST
@login_required
def archive_po(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)

    role = get_effective_role(request.user)
    if not (request.user.is_superuser or role in {"TopManagement"} or is_top_management(request.user)):
        raise PermissionDenied("Only Accounting Head / Top Management can archive POs.")

    po.is_archived = True
    po.status = "ARCHIVED"
    po.approval_status = POApprovalStatus.APPROVED
    po.save(update_fields=["is_archived", "status", "updated_at"])
    po.log_update(request.user, "Archived PO.")

    latest = po.updates.first()
    return JsonResponse({
        "ok": True,
        "system_message": latest.system_update if latest else "",
    })



@login_required
def dr_table_export(request):
    qs = DeliveryReceipt.objects.select_related(
        "client", "agent", "created_by", "source_dr"
    )
    # -------------------
    # SAME FILTERS AS dr_table (SMART, MULTI)
    # -------------------

    agent_ids = clean_int_list(request.GET.getlist("agent"))
    client_names = [c.strip() for c in request.GET.getlist("client_name") if c.strip()]
    dr_numbers = [d.strip() for d in request.GET.getlist("dr_number") if d.strip()]

    payment_methods = [p for p in request.GET.getlist("payment_method") if p]
    payment_statuses = [p for p in request.GET.getlist("payment_status") if p]
    delivery_statuses = [d for d in request.GET.getlist("delivery_status") if d]
    delivery_methods = [d for d in request.GET.getlist("delivery_method") if d]

    q = (request.GET.get("q") or "").strip()

    hide_archived = request.GET.get("hide_archived") == "1"
    hide_cancelled = request.GET.get("hide_cancelled") == "1"
    with_sales_invoice = request.GET.get("with_sales_invoice") == "1"

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    due_start = request.GET.get("due_start")
    due_end = request.GET.get("due_end")

    # ---- Visibility flags ----
    if hide_archived:
        qs = qs.filter(is_archived=False)

    if hide_cancelled:
        qs = qs.filter(is_cancelled=False)

    if with_sales_invoice:
        qs = qs.exclude(sales_invoice_no__isnull=True).exclude(sales_invoice_no="")

    # ---- Agent (multi) ----
    if agent_ids:
        qs = qs.filter(agent_id__in=agent_ids)

    # ---- Client name (multi, NAME-BASED ONLY) ----
    if client_names:
        q_obj = Q()
        for cname in client_names:
            q_obj |= Q(client__company_name__icontains=cname)
        qs = qs.filter(q_obj)

    # ---- DR numbers (multi) ----
    if dr_numbers:
        q_obj = Q()
        for drn in dr_numbers:
            q_obj |= Q(dr_number__icontains=drn)
        qs = qs.filter(q_obj)

    # ---- Payment / Delivery filters (multi) ----
    if payment_methods:
        qs = qs.filter(payment_method__in=payment_methods)

    if payment_statuses:
        qs = qs.filter(payment_status__in=payment_statuses)

    if delivery_statuses:
        qs = qs.filter(delivery_status__in=delivery_statuses)

    if delivery_methods:
        qs = qs.filter(delivery_method__in=delivery_methods)

    # ---- Date filters ----
    if start_date:
        qs = qs.filter(date_of_order__gte=start_date)

    if end_date:
        qs = qs.filter(date_of_order__lte=end_date)

    if due_start:
        qs = qs.filter(payment_due__gte=due_start)

    if due_end:
        qs = qs.filter(payment_due__lte=due_end)

    # ---- Free-text keyword (client OR agent) ----
    if q:
        qs = qs.filter(
            Q(client__company_name__icontains=q) |
            Q(agent__username__icontains=q) |
            Q(agent__first_name__icontains=q) |
            Q(agent__last_name__icontains=q)
        )

    SORT_OPTIONS = {
        "date_desc": "-date_of_order",
        "date_asc": "date_of_order",
        "total_desc": "-total_amount",
        "total_asc": "total_amount",
        "dr_desc": "-dr_number",
        "dr_asc": "dr_number",
    }

    sort_by = request.GET.get("sort_by", "dr_asc")
    qs = qs.order_by(SORT_OPTIONS.get(sort_by, "-date_of_order"))

    # ---- EXCEL ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delivery Receipts"

    headers = [
        "DR Number", "Client", "Agent",
        "Date Ordered", "Date Delivered",
        "Due Date", "Payment Due",
        "Delivery Status", "Payment Status",
        "Payment Method", "Delivery Method",
        "Total Amount",
        "Approval Status",
        "Remarks", "Payment Details",
        "Source DR",
        "Created By",
        "Created At", "Updated At",
        "Archived"
    ]
    ws.append(headers)

    for dr in qs:
        ws.append([
            dr.dr_number,
            dr.client.company_name,
            dr.agent.get_full_name() or dr.agent.username,
            dr.date_of_order,
            dr.date_of_delivery,
            dr.due_date,
            dr.payment_due,
            dr.get_delivery_status_display(),
            dr.get_payment_status_display(),
            dr.get_payment_method_display(),
            dr.get_delivery_method_display(),
            float(dr.total_amount),
            dr.get_approval_status_display(),
            dr.remarks,
            dr.payment_details,
            dr.source_dr.dr_number if dr.source_dr else "",
            dr.created_by.get_full_name() or dr.created_by.username,
            excel_safe_datetime(dr.created_at),   # ✅ FIX
            excel_safe_datetime(dr.updated_at),   # ✅ FIX
            "Yes" if dr.is_archived else "No",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=delivery_receipts.xlsx"
    wb.save(response)
    return response

@require_POST
@login_required
def po_submit(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    po.submit_to_next(request.user)
    return redirect("po-edit", pk=pk)

@require_POST
@login_required
def po_complete(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    po.submit_to_next(request.user)
    return redirect("po-edit", pk=pk)

@require_GET
@login_required
def po_filter_suggestions_api(request):
    """
    Global PO table smart-search suggestions (complete across ALL POs).
    """
    q = (request.GET.get("q") or "").strip()
    if not q:
        return JsonResponse({"ok": True, "results": []})

    q_lower = q.lower()
    results = []

    # ---- Prepared By (users) ----
    users_qs = (
        User.objects
        .filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q)
        )
        .order_by("username")[:10]
    )
    for u in users_qs:
        label = u.get_full_name() or u.username
        results.append({
            "type": "prepared_by",
            "badge": "Prepared By",
            "value": str(u.id),
            "label": label,
        })

    # ---- Product ID ----
    pid_qs = (
        ProductID.objects
        .filter(Q(code__icontains=q) | Q(description__icontains=q))
        .order_by("code")[:10]
    )
    for p in pid_qs:
        results.append({
            "type": "product_id",
            "badge": "Product ID",
            "value": str(p.id),
            "label": p.code,
        })

    # ---- Paid To (distinct values across all POs) ----
    paid_to_qs = (
        PurchaseOrder.objects
        .exclude(paid_to__isnull=True)
        .exclude(paid_to__exact="")
        .filter(paid_to__icontains=q)
        .values_list("paid_to", flat=True)
        .distinct()
        .order_by("paid_to")[:10]
    )
    for p in paid_to_qs:
        results.append({
            "type": "paid_to",
            "badge": "Paid To",
            "value": p,
            "label": p,
        })

    # ---- PO numbers ----
    po_nums = (
        PurchaseOrder.objects
        .exclude(po_number__isnull=True)
        .exclude(po_number__exact="")
        .filter(po_number__icontains=q)
        .values_list("po_number", flat=True)
        .distinct()
        .order_by("po_number")[:10]
    )
    for n in po_nums:
        results.append({
            "type": "po_number",
            "badge": "PO #",
            "value": n,
            "label": n,
        })

    # ---- Statuses (humanized) ----
    for val, label in POStatus.choices:
        if q_lower in val.lower() or q_lower in label.lower():
            results.append({
                "type": "status",
                "badge": "Status",
                "value": val,
                "label": label,
            })

    for val, label in POApprovalStatus.choices:
        if q_lower in val.lower() or q_lower in label.lower():
            results.append({
                "type": "approval_status",
                "badge": "Approval",
                "value": val,
                "label": label,
            })

    # ---- Free text fallback ----
    results.append({
        "type": "q",
        "badge": "Keyword",
        "value": q,
        "label": f'Search "{q}" (Paid To or Prepared By)',
    })

    return JsonResponse({"ok": True, "results": results[:25]})

@login_required
def po_table(request):
    qs = (
        PurchaseOrder.objects
        .select_related("prepared_by")
    )



    # -------------------
    # Filters
    # -------------------
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    paid_to_list = [x.strip() for x in request.GET.getlist("paid_to") if x.strip()]
    prepared_by_ids = clean_int_list(request.GET.getlist("prepared_by"))
    status_list = [x for x in request.GET.getlist("status") if x]
    approval_status_list = [x for x in request.GET.getlist("approval_status") if x]
    product_id_ids = clean_int_list(request.GET.getlist("product_id"))

    q = (request.GET.get("q") or "").strip()

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    sort_by = request.GET.get("sort_by", "po_asc")

    hide_archived = request.GET.get("hide_archived") == "1"
    hide_cancelled = request.GET.get("hide_cancelled") == "1"


    # Paid To (multi icontains OR)
    if paid_to_list:
        q_obj = Q()
        for p in paid_to_list:
            q_obj |= Q(paid_to__icontains=p)
        qs = qs.filter(q_obj)

    # Prepared By (multi)
    if prepared_by_ids:
        qs = qs.filter(prepared_by_id__in=prepared_by_ids)

    # Product ID (multi)
    if product_id_ids:
        qs = qs.filter(product_id_ref_id__in=product_id_ids)

    # Status (multi)
    if status_list:
        qs = qs.filter(status__in=status_list)

    # Approval Status (multi)
    if approval_status_list:
        qs = qs.filter(approval_status__in=approval_status_list)

    # Date range
    if start_date:
        qs = qs.filter(date__gte=start_date)
    if end_date:
        qs = qs.filter(date__lte=end_date)

    # Free text keyword (Paid To OR Prepared By)
    if q:
        qs = qs.filter(
            Q(paid_to__icontains=q) |
            Q(prepared_by__username__icontains=q) |
            Q(prepared_by__first_name__icontains=q) |
            Q(prepared_by__last_name__icontains=q)
        )


    if start_date:
        qs = qs.filter(date__gte=start_date)

    if end_date:
        qs = qs.filter(date__lte=end_date)
    if hide_archived and hide_cancelled:
        # SHOW ALL: active + archived + cancelled
        pass

    elif hide_cancelled:
        # Show cancelled (which are archived) + all active
        qs = qs.filter(
            Q(is_cancelled=True) |
            Q(is_archived=False)
        )

    elif hide_archived:
        # Show archived BUT NOT cancelled, plus active
        qs = qs.filter(
            Q(is_archived=False) |
            Q(is_archived=True, is_cancelled=False)
        )

    else:
        # Default: show active only
        qs = qs.filter(
            is_archived=False,
            is_cancelled=False
        )

    SORT_OPTIONS = {
        "date_desc": "-date",
        "date_asc": "date",
        "total_desc": "-total",
        "total_asc": "total",
        "po_desc": "-po_number",
        "po_asc": "po_number",
    }
    qs = qs.order_by(SORT_OPTIONS.get(sort_by, "-date"))
    # -------------------------------------------------
    # Archived / Cancelled visibility logic
    # -------------------------------------------------


    # -------------------
    # Pagination (LAST)
    # -------------------
    paginator = Paginator(qs, 100)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    paid_to_values = (
        PurchaseOrder.objects
        .exclude(paid_to__isnull=True)
        .exclude(paid_to__exact="")
        .values_list("paid_to", flat=True)
        .distinct()
        .order_by("paid_to")
    )
    total_sum = qs.aggregate(total=Sum("total"))["total"] or 0
    # -----------------------------
    # Pagination-safe querystring
    # -----------------------------
    qs = request.GET.copy()
    qs.pop("page", None)
    base_querystring = qs.urlencode()

    return render(request, "bondking_app/po_table.html", {
        "page_obj": page_obj,
        "users": User.objects.all().order_by("username"),
        "approval_statuses": [
            ("PENDING", "Pending"),
            ("APPROVED", "Approved"),
            ("DECLINED", "Declined"),
        ],
        "hide_archived": hide_archived,
        "hide_cancelled": hide_cancelled,
        "sort_by": sort_by,
        "selected": {
            "paid_to": paid_to_list,
            "prepared_by": prepared_by_ids,
            "status": status_list,
            "approval_status": approval_status_list,
            "product_id": product_id_ids,
            "start_date": start_date,
            "end_date": end_date,
            "q": q,
        },
        "total_sum": total_sum,
        "paid_to_values": paid_to_values,
        "is_top_management": is_top_management(request.user),
        "product_ids": ProductID.objects.filter(is_active=True).order_by("code"),
        "statuses": [s for s, _ in POStatus.choices],
        "product_id": product_id_ids,
        "base_querystring": base_querystring,
    })


@login_required
def po_table_export(request):
    from datetime import datetime
    from django.db.models import Q, Sum
    import openpyxl
    from django.http import HttpResponse

    qs = PurchaseOrder.objects.select_related("prepared_by")

    # -------------------
    # Excel-safe datetime
    # -------------------
    def excel_safe_datetime(value):
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)
        return value

    # -------------------
    # READ FILTERS (MULTI)
    # -------------------
    paid_to_list = [v.strip() for v in request.GET.getlist("paid_to") if v.strip()]
    prepared_by_ids = [
        int(v) for v in request.GET.getlist("prepared_by") if v.isdigit()
    ]
    status_list = [v for v in request.GET.getlist("status") if v]
    approval_status_list = [v for v in request.GET.getlist("approval_status") if v]
    product_id_ids = [
        int(v) for v in request.GET.getlist("product_id") if v.isdigit()
    ]

    q = (request.GET.get("q") or "").strip()

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    sort_by = request.GET.get("sort_by", "po_asc")

    hide_archived = request.GET.get("hide_archived") == "1"
    hide_cancelled = request.GET.get("hide_cancelled") == "1"

    # -------------------
    # APPLY FILTERS
    # -------------------

    # Paid To (OR across tags)
    if paid_to_list:
        q_obj = Q()
        for p in paid_to_list:
            q_obj |= Q(paid_to__icontains=p)
        qs = qs.filter(q_obj)

    # Prepared By
    if prepared_by_ids:
        qs = qs.filter(prepared_by_id__in=prepared_by_ids)

    # Product ID
    if product_id_ids:
        qs = qs.filter(product_id_ref_id__in=product_id_ids)

    # Status
    if status_list:
        qs = qs.filter(status__in=status_list)

    # Approval Status
    if approval_status_list:
        qs = qs.filter(approval_status__in=approval_status_list)

    # Date range
    if start_date:
        qs = qs.filter(date__gte=start_date)
    if end_date:
        qs = qs.filter(date__lte=end_date)

    # Free text keyword
    if q:
        qs = qs.filter(
            Q(paid_to__icontains=q) |
            Q(prepared_by__username__icontains=q) |
            Q(prepared_by__first_name__icontains=q) |
            Q(prepared_by__last_name__icontains=q)
        )

    if hide_archived and hide_cancelled:
        # SHOW ALL: active + archived + cancelled
        pass

    elif hide_cancelled:
        # Show cancelled (which are archived) + all active
        qs = qs.filter(
            Q(is_cancelled=True) |
            Q(is_archived=False)
        )

    elif hide_archived:
        # Show archived BUT NOT cancelled, plus active
        qs = qs.filter(
            Q(is_archived=False) |
            Q(is_archived=True, is_cancelled=False)
        )

    else:
        # Default: show active only
        qs = qs.filter(
            is_archived=False,
            is_cancelled=False
        )


    # -------------------
    # SORT (MATCH po_table)
    # -------------------
    SORT_MAP = {
        "date_desc": "-date",
        "date_asc": "date",
        "total_desc": "-total",
        "total_asc": "total",
        "po_desc": "-po_number",
        "po_asc": "po_number",
    }
    qs = qs.order_by(SORT_MAP.get(sort_by, "po_number"))

    # -------------------
    # EXCEL OUTPUT
    # -------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    headers = [
        "PO Number",
        "Date",
        "Paid To",
        "Address",
        "Prepared By",
        "Status",
        "Approval Status",
        "Total",
        "Total Billed",
        "Balance",
        "Created At",
        "Updated At",
        "Archived",
        "Cancelled",
    ]
    ws.append(headers)

    for po in qs:
        ws.append([
            po.po_number,
            po.date,
            po.paid_to,
            po.address,
            po.prepared_by.get_full_name() or po.prepared_by.username,
            po.status,
            po.approval_status,
            float(po.total or 0),
            float(po.billed_total() or 0),
            float(po.balance_amount() or 0),
            excel_safe_datetime(po.created_at),
            excel_safe_datetime(po.updated_at),
            "Yes" if po.is_archived else "No",
            "Yes" if po.is_cancelled else "No",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=purchase_orders.xlsx"
    wb.save(response)

    return response



@require_POST
@login_required
def cancel_dr(request, pk):
    if not is_top_management(request.user):
        raise PermissionDenied("Only Top Management can cancel DRs.")

    dr = get_object_or_404(DeliveryReceipt, pk=pk)

    dr.is_cancelled = True
    dr.is_archived = True   # hides from kanban
    dr.save(update_fields=["is_cancelled", "is_archived", "updated_at"])

    dr.log_update(request.user, "Cancelled DR.")
    return JsonResponse({"ok": True})

@require_POST
@login_required
def cancel_po(request, pk):
    if not can_cancel_po(request.user):
        raise PermissionDenied("Only RVT can cancel POs.")

    po = get_object_or_404(PurchaseOrder, pk=pk)

    po.is_cancelled = True
    po.is_archived = True
    po.status = POStatus.ARCHIVED
    po.save(update_fields=["is_cancelled", "is_archived", "status", "updated_at"])

    # cancel all billings (not delete)
    po.billings.update(is_cancelled=True)

    po.log_update(request.user, "Cancelled PO (and cancelled all billings).")
    return redirect("po-table")


def compute_stock_snapshot():
    snapshot = []

    for product in Product.objects.all():
        wh_in = InventoryIssuanceItem.objects.filter(
            issuance__issuance_type="TF_TO_WH",
            issuance__is_pending=False,
            issuance__is_cancelled=False,
            product=product,
        ).aggregate(q=Sum("quantity"))["q"] or 0

        wh_out = InventoryIssuanceItem.objects.filter(
            issuance__issuance_type="WH_TO_HQ",
            issuance__is_pending=False,
            issuance__is_cancelled=False,
            product=product,
        ).aggregate(q=Sum("quantity"))["q"] or 0

        dr_out = DeliveryReceiptItem.objects.filter(
            product=product,
        ).filter(
            Q(delivery_receipt__delivery_status="DELIVERED") |
            Q(delivery_receipt__delivery_method=DeliveryMethod.D2D_STOCKS)
        ).exclude(
            delivery_receipt__delivery_method=DeliveryMethod.DOOR_TO_DOOR
        ).exclude(
            delivery_receipt__is_cancelled=True
        ).aggregate(q=Sum("quantity"))["q"] or 0


        snapshot.append({
            "product": product,
            "wh_stock": wh_in - wh_out,
            "hq_stock": wh_out - dr_out,
        })

    return snapshot

@login_required
def inventory_table(request):
    user = request.user
    is_top_management = user.groups.filter(name="TopManagement").exists() or user.is_superuser
    is_logistics = user.groups.filter(name__in=["LogisticsOfficer", "LogisticsHead"]).exists()
    is_logistics_head = user.groups.filter(name="LogisticsHead").exists()
    can_approve_inventory = is_logistics_head or is_top_management


    # ==========================
    # SNAPSHOT (ALL TIME, UNFILTERED)
    # ==========================
    snapshot = compute_stock_snapshot()
    # ✅ ALWAYS define products
    products = Product.objects.all()

    # ==========================
    # FILTER PARAMS
    # ==========================
    selected_types = request.GET.getlist("type")
    selected_products = request.GET.getlist("product")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    sort_by = request.GET.get("sort_by", "date_desc")
    hide_cancelled = request.GET.get("hide_cancelled") == "1"

    # ==========================
    # UNIFIED MOVEMENT ROWS
    # ==========================
    rows = []

    issuance_items = InventoryIssuanceItem.objects.select_related(
        "issuance", "product"
    )

    for item in issuance_items:
        iss = item.issuance

        if hide_cancelled and iss.is_cancelled:
            continue


        if selected_types and iss.issuance_type not in selected_types:
            continue

        if selected_products and str(item.product.id) not in selected_products:
            continue

        rows.append({
            "date": iss.created_at,
            "type": iss.issuance_type,
            "type_label": iss.get_issuance_type_display(),
            "ref": f"ISS-{iss.id}",
            "group_key": f"ISS-{iss.id}",
            "parent_id": iss.id,
            "parent_type": "ISSUANCE",
            "product": item.product,
            "qty": item.quantity,
            "from": "TF" if iss.issuance_type == InventoryIssuance.TF_TO_WH else "WH",
            "to": "WH" if iss.issuance_type == InventoryIssuance.TF_TO_WH else "HQ",
            "is_pending": iss.is_pending,
            "is_cancelled": iss.is_cancelled,
        })

    dr_items = DeliveryReceiptItem.objects.filter(
        Q(delivery_receipt__delivery_method=DeliveryMethod.D2D_STOCKS) |
        Q(delivery_receipt__delivery_status="DELIVERED")
    ).exclude(
        delivery_receipt__delivery_method=DeliveryMethod.DOOR_TO_DOOR
    )


    if not hide_cancelled:
        dr_items = dr_items.filter(delivery_receipt__is_cancelled=True)


    for item in dr_items:
        dr = item.delivery_receipt

        if selected_types and "DR" not in selected_types:
            continue

        if selected_products and str(item.product.id) not in selected_products:
            continue

        rows.append({
            "date": dr.date_of_delivery,
            "type": "DR",
            "type_label": "Delivery Receipt",
            "ref": dr.dr_number,
            "group_key": f"DR-{dr.id}",
            "parent_id": dr.id,
            "parent_type": "DR",
            "product": item.product,
            "qty": item.quantity,
            "from": "HQ",
            "to": dr.client.company_name,
            "to_client_id": dr.client.id,
            "to_client_name": dr.client.company_name,
            "is_pending": False,
            "is_cancelled": dr.is_cancelled,
        })

    # ==========================
    # DATE FILTER
    # ==========================

    if start_date:
        start = datetime.fromisoformat(start_date).date()
        rows = [
            r for r in rows
            if r["date"] and r["date"].date() >= start
        ]

    if end_date:
        end = datetime.fromisoformat(end_date).date()
        rows = [
            r for r in rows
            if r["date"] and r["date"].date() <= end
        ]

    # ==========================
    # SORTING
    # ==========================
    reverse = sort_by.endswith("desc")
    def normalize_date(d):
        if d is None:
            return date.min
        if isinstance(d, datetime):
            return d.date()
        return d  # already a date

    rows.sort(
        key=lambda r: normalize_date(r["date"]),
        reverse=reverse
    )


    # ==========================
    # EXPORT
    # ==========================
    if "export" in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Date",
            "Movement",
            "Reference",
            "Item",
            "Quantity",
            "From",
            "To",
            "Client",
            "Status",
        ])
        def excel_safe_date(value):
            if isinstance(value, datetime):
                return value.replace(tzinfo=None)
            return value

        for r in rows:
            ws.append([
                excel_safe_date(r["date"]),
                r["type_label"],
                r["ref"],
                r["product"].name,
                r["qty"],
                r["from"],
                r["to"],
                r.get("to_client_name", "") if r["parent_type"] == "DR" else "",
                "Pending" if r["is_pending"] else
                "Cancelled" if r["is_cancelled"] else
                "Approved",
            ])


        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=inventory.xlsx"
        wb.save(response)
        return response
    nav_ids = [
        str(r["parent_id"])
        for r in rows
        if r.get("parent_type") == "ISSUANCE"
    ]


    return render(request, "bondking_app/inventory_table.html", {
        "snapshot": snapshot,
        "rows": rows,
        "products": products,
        "selected": {
            "types": selected_types,
            "products": selected_products,
            "start_date": start_date,
            "end_date": end_date,
        },
        "sort_by": sort_by,
        "hide_cancelled": hide_cancelled,
        "is_top_management": is_top_management,
        "is_logistics": is_logistics,
        "can_approve_inventory": can_approve_inventory,
        "can_manage_inventory_issuance": can_manage_inventory_issuance(user),
        "nav_ids": ",".join(nav_ids),
    })
@login_required
def inventory_cancel(request, pk):
    issuance = get_object_or_404(InventoryIssuance, pk=pk)
    if not can_manage_inventory_issuance(request.user):
        raise PermissionDenied("You are not allowed to cancel this issuance.")

    issuance.is_cancelled = True
    issuance.is_pending = False
    issuance.save(update_fields=["is_cancelled", "is_pending"])

    return JsonResponse({
        "ok": True,
        "message": "Inventory issuance cancelled.",
        "redirect": reverse("inventory-table"),
        })


@login_required
def inventory_new(request):
    user = request.user
    is_locked = False  # NEW form is editable

    # Permission check (same logic as inventory table)
    if not can_manage_inventory_issuance(user):
        messages.error(request, "Only AGR is allowed to create inventory issuances.")
        return redirect("inventory-table")  
    # ==========================
    # WH STOCK MAP (for display)
    # ==========================
    wh_stock_map = {}

    for product in Product.objects.all():
        wh_in = InventoryIssuanceItem.objects.filter(
            issuance__issuance_type=InventoryIssuance.TF_TO_WH,
            issuance__is_pending=False,
            issuance__is_cancelled=False,
            product=product,
        ).aggregate(q=Sum("quantity"))["q"] or 0

        wh_out = InventoryIssuanceItem.objects.filter(
            issuance__issuance_type=InventoryIssuance.WH_TO_HQ,
            issuance__is_pending=False,
            issuance__is_cancelled=False,
            product=product,
        ).aggregate(q=Sum("quantity"))["q"] or 0

        wh_stock_map[product.id] = wh_in - wh_out

    if request.method == "POST":
        form = InventoryIssuanceForm(request.POST)
        formset = InventoryIssuanceItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            issuance_type = form.cleaned_data["issuance_type"]

            # ==========================
            # WH AVAILABILITY VALIDATION
            # ==========================
            if issuance_type == InventoryIssuance.WH_TO_HQ:
                for f in formset:
                    product = f.cleaned_data.get("product")
                    qty = f.cleaned_data.get("quantity")

                    if not product or not qty:
                        continue

                    # WH stock computation
                    wh_in = InventoryIssuanceItem.objects.filter(
                        issuance__issuance_type=InventoryIssuance.TF_TO_WH,
                        issuance__is_pending=False,
                        issuance__is_cancelled=False,
                        product=product,
                    ).aggregate(q=Sum("quantity"))["q"] or 0

                    wh_out = InventoryIssuanceItem.objects.filter(
                        issuance__issuance_type=InventoryIssuance.WH_TO_HQ,
                        issuance__is_pending=False,
                        issuance__is_cancelled=False,
                        product=product,
                    ).aggregate(q=Sum("quantity"))["q"] or 0

                    available_wh = wh_in - wh_out

                    if qty > available_wh:
                        messages.error(
                            request,
                            f"Not enough WH stock for {product.name}. "
                            f"Available: {available_wh}, Requested: {qty}"
                        )
                        return redirect("inventory-new")

            # ==========================
            # SAVE (ATOMIC)
            # ==========================
            with transaction.atomic():
                issuance = form.save(commit=False)
                issuance.created_by = user
                issuance.is_pending = False
                issuance.save()

                formset.instance = issuance
                formset.save()

            messages.success(request, "Inventory issuance created successfully.")
            return redirect("inventory-table")
        else:
            messages.error(request, "Please correct the errors below.")

    else:
        form = InventoryIssuanceForm()
        formset = InventoryIssuanceItemFormSet(prefix="items")

    return render(request, "bondking_app/inventory_form.html", {
        "form": form,
        "formset": formset,
        "wh_stock_map": wh_stock_map,
        "is_locked": is_locked,
    })
@login_required
def inventory_edit(request, pk):
    issuance = get_object_or_404(InventoryIssuance, pk=pk)
    is_locked = issuance.is_cancelled

    if not can_manage_inventory_issuance(request.user):
        is_locked = True


    # reuse the same form & formset logic as inventory_new
    form = InventoryIssuanceForm(instance=issuance)
    formset = InventoryIssuanceItemFormSet(
        request.POST or None,
        instance=issuance,
        form_kwargs={"is_locked": is_locked},
    )

    if request.method == "POST" and not is_locked:
        form = InventoryIssuanceForm(request.POST, instance=issuance)
        formset = InventoryIssuanceItemFormSet(
            request.POST or None,
            instance=issuance,
            form_kwargs={"is_locked": is_locked},
        )

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect("inventory-table")
    # =========================
    # NAVIGATION (DR-style)
    # =========================
    prev_issuance = next_issuance = None
    nav_querystring = ""

    # ========= MODE 1: nav_ids (Bulk Open) =========
    nav_ids = request.GET.get("nav_ids")
    if nav_ids:
        ids = [int(i) for i in nav_ids.split(",") if i.isdigit()]
        if pk in ids:
            idx = ids.index(pk)
            if idx > 0:
                prev_issuance = InventoryIssuance.objects.filter(pk=ids[idx - 1]).first()
            if idx < len(ids) - 1:
                next_issuance = InventoryIssuance.objects.filter(pk=ids[idx + 1]).first()

        nav_querystring = f"nav_ids={nav_ids}"

    # ========= MODE 2: filter-based (Reference click) =========
    elif request.GET.get("from") == "table":
        qs = InventoryIssuance.objects.all()

        # APPLY SAME FILTERS AS TABLE
        if request.GET.getlist("type"):
            qs = qs.filter(issuance_type__in=request.GET.getlist("type"))
        if request.GET.getlist("product"):
            qs = qs.filter(items__product_id__in=request.GET.getlist("product")).distinct()
        if request.GET.get("start_date"):
            qs = qs.filter(date__gte=request.GET["start_date"])
        if request.GET.get("end_date"):
            qs = qs.filter(date__lte=request.GET["end_date"])

        ids = list(qs.order_by("-date").values_list("id", flat=True))

        if pk in ids:
            idx = ids.index(pk)
            if idx > 0:
                prev_issuance = InventoryIssuance.objects.filter(pk=ids[idx - 1]).first()
            if idx < len(ids) - 1:
                next_issuance = InventoryIssuance.objects.filter(pk=ids[idx + 1]).first()

        nav_querystring = request.GET.urlencode()

    return render(request, "bondking_app/inventory_form.html", {
        "form": form,
        "formset": formset,
        "issuance": issuance,
        "is_locked": is_locked,
        "prev_issuance": prev_issuance,
        "next_issuance": next_issuance,
        "nav_querystring": nav_querystring,
        "can_manage_inventory_issuance": can_manage_inventory_issuance(request.user),
    })

@login_required
@require_POST
def inventory_delete(request, pk):
    issuance = get_object_or_404(InventoryIssuance, pk=pk)

    # Permission check
    if not can_manage_inventory_issuance(request.user):
        return JsonResponse({
            "ok": False,
            "error": "You do not have permission to delete this issuance."
        }, status=403)

    # 🔐 Atomic delete
    with transaction.atomic():
        # Delete child items first (explicit & safe)
        InventoryIssuanceItem.objects.filter(issuance=issuance).delete()

        # Delete parent issuance
        issuance.delete()

    return JsonResponse({
        "ok": True,
        "message": "Inventory issuance deleted successfully.",
        "redirect": reverse("inventory-table"),
    })


@require_POST
@login_required
def product_id_quick_create(request):
    code = request.POST.get("code")
    description = request.POST.get("description", "")

    if not code:
        return JsonResponse({"ok": False, "error": "Product ID code is required."})

    obj, created = ProductID.objects.get_or_create(
        code=code.strip(),
        defaults={"description": description},
    )

    return JsonResponse({
        "ok": True,
        "id": obj.id,
        "label": f"{obj.code}",
    })





@login_required
def dr_print(request, pk):
    dr = get_object_or_404(DeliveryReceipt, pk=pk)

    html = render_to_string(
        "bondking_app/dr_print.html",
        {
            "dr": dr,
            "client": dr.client,
            "items": dr.items.select_related("product").all(),
            "shipping": 0,
            "other": 0,
        },
        request=request,  # ✅ IMPORTANT
    )
    # ✅ Explicit environment detection
    IS_RENDER = os.environ.get("RENDER") == "true"

    if IS_RENDER:
        # ===== PRODUCTION (Render / Linux) =====
        from weasyprint import HTML

        pdf = HTML(
            string=html,
            base_url=settings.STATIC_ROOT  # 🔑 THIS IS THE FIX
        ).write_pdf()

    else:
        # ===== LOCAL (Windows) =====
        import pdfkit

        config = pdfkit.configuration(
            wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
        )

        pdf = pdfkit.from_string(
            html,
            False,
            configuration=config,
            options={
                "page-size": "Letter",
                "orientation": "Portrait",
                "margin-top": "15mm",
                "margin-bottom": "15mm",
                "margin-left": "15mm",
                "margin-right": "15mm",
                "encoding": "UTF-8",
                "--enable-local-file-access": ""
            }
        )

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="DR-{dr.dr_number}.pdf"'
    return response

@login_required
def po_print(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)

    html = render_to_string(
        "bondking_app/po_print.html",
        {
            "po": po,
            "items": po.particulars.all(),
        },
        request=request,  # ✅ REQUIRED (matches dr_print)
    )

    # ✅ Explicit environment detection (MATCH dr_print)
    IS_RENDER = os.environ.get("RENDER") == "true"

    if IS_RENDER:
        # ===== PRODUCTION (Render / Linux) =====
        from weasyprint import HTML

        pdf = HTML(
            string=html,
            base_url=settings.STATIC_ROOT  # 🔑 REQUIRED FOR STATIC FILES
        ).write_pdf()

    else:
        # ===== LOCAL (Windows) =====
        config = pdfkit.configuration(
            wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
        )

        pdf = pdfkit.from_string(
            html,
            False,
            configuration=config,
            options={
                "page-size": "Letter",
                "orientation": "Portrait",
                "margin-top": "15mm",
                "margin-bottom": "15mm",
                "margin-left": "15mm",
                "margin-right": "15mm",
                "encoding": "UTF-8",
                "--enable-local-file-access": ""
            }
        )

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="PO-{po.po_number}.pdf"'
    return response

@login_required
def client_table(request):
    qs = Client.objects.all()

    # -------------------
    # Filters
    # -------------------
    company = request.GET.get("company", "")
    agent = request.GET.get("agent", "")
    city = request.GET.get("city", "")

    sort_by = request.GET.get("sort_by", "company_asc")
    companies = (
        Client.objects
        .exclude(company_name__isnull=True)
        .exclude(company_name__exact="")
        .values_list("company_name", flat=True)
        .distinct()
        .order_by("company_name")
    )

    agents = (
        User.objects
        .filter(clients__isnull=False)
        .distinct()
        .order_by("username")
    )

    cities = (
        Client.objects
        .exclude(city_municipality__isnull=True)
        .exclude(city_municipality__exact="")
        .values_list("city_municipality", flat=True)
        .distinct()
        .order_by("city_municipality")
    )

    if company:
        qs = qs.filter(company_name__icontains=company)

    if agent.isdigit():
        qs = qs.filter(agent_id=int(agent))

    if city:
        qs = qs.filter(city_municipality__icontains=city)


    SORT_OPTIONS = {
        "company_asc": "company_name",
        "company_desc": "-company_name",
        "created_desc": "-created_at",
        "created_asc": "created_at",
    }
    qs = qs.order_by(SORT_OPTIONS.get(sort_by, "company_name"))

    # -------------------
    # Pagination
    # -------------------
    paginator = Paginator(qs, 100)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "bondking_app/client_table.html", {
        "page_obj": page_obj,
        "sort_by": sort_by,
        "selected": {
            "company": company,
            "agent": agent,
            "city": city,
        },
        "suggestions": {
            "companies": companies,
            "agents": agents,
            "cities": cities,
        },
    })


@login_required
def client_edit(request, pk):
    client = get_object_or_404(Client, pk=pk)

    if request.method == "POST":
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect("client-table")
    else:
        form = ClientForm(instance=client)

    return render(request, "bondking_app/client_form.html", {
        "form": form,
        "client": client,
        "is_create": False,
    })

@require_POST
@login_required
def billing_advance(request, pk):
    billing = get_object_or_404(Billing, pk=pk)

    po = billing.source_po
    if po.is_archived or po.status == POStatus.ARCHIVED:
        return JsonResponse({"ok": False, "error": "PO is archived."}, status=400)

    # Only allow while PO is in BILLING stage
    if po.status != POStatus.BILLING:
        return JsonResponse({"ok": False, "error": "PO is not in Billing stage."}, status=400)

    role = get_user_role(request.user)
    # =========================
    # SAVE PROOF OF PAYMENT IF SENT
    # =========================
    if "proof_of_payment" in request.FILES:
        billing.proof_of_payment = request.FILES["proof_of_payment"]
        billing.save(update_fields=["proof_of_payment"])
    # Status progression
    order = [
        BillingStatus.CHECK_CREATION,
        BillingStatus.CHECK_SIGNING,
        BillingStatus.PAYMENT_RELEASE,
        BillingStatus.PAID,
    ]
    idx = order.index(billing.status) if billing.status in order else 0
    current = order[idx]
    # =========================
    # ENFORCE PROOF OF PAYMENT
    # =========================
    next_status = order[idx + 1] if idx + 1 < len(order) else None

    if next_status == BillingStatus.PAID:
        if not billing.proof_of_payment:
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Proof of payment is required before releasing payment."
                },
                status=400
            )

    # Who can advance each step
    allowed = {
        BillingStatus.CHECK_CREATION: {"RVT", "AccountingOfficer", "AccountingHead"},
        BillingStatus.CHECK_SIGNING: {"AGR"},
        BillingStatus.PAYMENT_RELEASE: {"RVT"},
        BillingStatus.PAID: set(),
    }

    if request.user.is_superuser:
        pass
    else:
        if role not in allowed.get(current, set()):
            return JsonResponse({"ok": False, "error": "You are not allowed to proceed this billing."}, status=403)

    if billing.is_cancelled:
        return JsonResponse({"ok": False, "error": "Billing is cancelled."}, status=400)

    if current == BillingStatus.PAID:
        return JsonResponse({"ok": True, "status": billing.status, "label": billing.get_status_display()})

    billing.status = order[idx + 1]
    billing.save(update_fields=["status"])

    po.log_update(request.user, f"Billing {billing.billing_number} advanced to {billing.get_status_display()}.")

    return JsonResponse({"ok": True, "status": billing.status, "label": billing.get_status_display()})


@require_POST
@login_required
def billing_cancel(request, pk):
    billing = get_object_or_404(Billing, pk=pk)

    if not can_cancel_po(request.user):
        return JsonResponse({"ok": False, "error": "Only RVT can cancel billings."}, status=403)

    if billing.is_cancelled:
        return JsonResponse({"ok": True})

    billing.is_cancelled = True
    billing.save(update_fields=["is_cancelled"])

    po = billing.source_po
    po.log_update(request.user, f"Billing {billing.billing_number} was cancelled.")

    return JsonResponse({"ok": True})



def client_table_export(request):
    # 🔁 Reuse the SAME filtering logic as client_table
    qs = Client.objects.all().select_related("agent")

    company = request.GET.get("company")
    agent = request.GET.get("agent")
    city = request.GET.get("city")

    if company:
        qs = qs.filter(company_name__icontains=company)

    if agent:
        qs = qs.filter(agent_id=agent)

    if city:
        qs = qs.filter(full_address__icontains=city)

    # =========================
    # Excel generation
    # =========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"

    ws.append([
        "Company",
        "Owner",
        "Contact",
        "Address",
        "Agent",
    ])

    for c in qs:
        ws.append([
            c.company_name,
            c.name_of_owner,
            c.contact_number,
            c.full_address,
            c.agent.get_full_name() if c.agent else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=clients.xlsx"
    wb.save(response)

    return response
